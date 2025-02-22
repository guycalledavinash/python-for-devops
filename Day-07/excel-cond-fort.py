### This program applies conditional formatting to an Excel file using openpyxl. It highlights cells based on their values.


from openpyxl import Workbook
from openpyxl.styles import PatternFill
from random import randint

wb = Workbook()
ws = wb.active

for row in range(2, 12):
    ws.append([randint(10, 100)])

for row in ws.iter_rows(min_row=2, max_row=11, min_col=1, max_col=1):
    for cell in row:
        if cell.value > 80:
            cell.fill = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
        elif cell.value > 50:
            cell.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
        else:
            cell.fill = PatternFill(start_color='99FF99', end_color='99FF99', fill_type='solid')

wb.save("formatted.xlsx")
